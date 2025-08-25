import os
import openpyxl
from openpyxl.workbook import Workbook
import pdfplumber
import re
from datetime import datetime

def main():
    try:
        directory = "pdf"
        files = os.listdir(directory)
        files_quantity = len(files)
    except Exception as error:
        print(f"\nException: {error}\nResult: Could not create directory with pdfs")

    if files_quantity == 0:
        raise Exception("\nException: No files found in the directory")

    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Invoice Imports"

        # Referenciando as colunas do Excel
        ws['A1'] = "Invoice #"
        ws['B1'] = "Date"
        ws['C1'] = "File name"
        ws['D1'] = "Status"

        last_empty_line = 1
        while ws["D" + str(last_empty_line)].value is not None:
            last_empty_line += 1
    except Exception as error:
        print(f"\nException: {error}\nResult: Could not create excel file")


    # Para file
    for file in files:
        try:
            # Abrir seu pdf
            with pdfplumber.open(directory + "/" + file) as pdf:
                first_page = pdf.pages[0]
                pdf_text = first_page.extract_text()

            # Instrucao Regex
            inv_number_re_pattern = r"INVOICE #(\d+)"
            inv_date_re_pattern = r"DATE: (\d{2}/\d{2}/\d{4})"

            # Procure o texto dentro do pdf baseado nessa instrucao regex
            match_number = re.search(inv_number_re_pattern, pdf_text)
            match_date = re.search(inv_date_re_pattern, pdf_text)

            # Se encontrou algo, ou seja, True :
            if match_number:
                # Pega o primeiro resultado e coloca dentro do Excel
                ws["A" + str(last_empty_line)] = match_number.group(1)
            else:
                raise Exception(f"Could not find invoice number")

            if match_date:
                ws["B" + str(last_empty_line)] = match_date.group(1)
            else:
                raise Exception(f"Could not find invoice number")

            ws["C" + str(last_empty_line)] = file
            ws["D" + str(last_empty_line)] = "Completed"

            last_empty_line += 1
        except Exception as error:
            print(f"\nException: {error}\nResult: Could not build the excel file")
            ws["C" + str(last_empty_line)] = file
            ws["D" + str(last_empty_line)] = f"Exception: {error}"
            last_empty_line += 1

    try:
        full_now = str(datetime.now()).replace(":", "-")
        dot_index = full_now.index(".")
        now = full_now[:dot_index]
        wb.save("Invoices -" + str(now) + ".xlsx")
    except Exception as error:
        print(f"\nException: {error}\nResult: Could not save the excel file")

if __name__ == "__main__":
    main()